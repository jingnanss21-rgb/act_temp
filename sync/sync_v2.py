"""
sync_v2.py - V2 每日活动快照同步脚本
独立于 sync_all.py，复用 doc_reader.py 和 supabase_writer.py

用法:
  python3 sync_v2.py              # 同步所有日期（按日期自动分流：>=2026-05-26 走 iWiki，否则走腾讯文档）
  python3 sync_v2.py --latest     # 只同步最新一天
  python3 sync_v2.py --date 0420  # 腾讯文档源：匹配导览文本中包含该串的条目
                                  # iWiki 源 + YYYY-MM-DD 写法：直接定位某天

数据源:
  - 旧（report_date < 2026-05-26）: 腾讯文档导览表 DWWhYU3FrWXhuSHhu → A列超链接 → 日表
  - 新（report_date >= 2026-05-26）: iWiki 页面 4020417529 → MMDD_餐饮活动日报.csv

目标表: tem_activity_daily (activity_id + report_date 为主键)
"""
import os
import re
import argparse
from datetime import date as _date_cls, datetime
from dotenv import load_dotenv

load_dotenv()

from doc_reader import (
    read_index_table_with_links,
    read_target_sheet,
    export_file,
    download_xlsx,
    extract_hyperlinks,
    extract_file_id_from_url,
)
from supabase_writer import get_client, upsert_batch
from iwiki_activity_loader import (
    load_iwiki_activity_attachments,
    load_activity_daily_from_iwiki,
)


# 旧腾讯文档导览表 file_id（< 2026-05-26 的历史日表都走这里）
NAV_FILE_ID = "DWWhYU3FrWXhuSHhu"

# 切换日期：从这一天起走 iWiki CSV 数据源
IWIKI_CUTOFF_DATE = _date_cls(2026, 5, 26)

# 底表字段 → DB 列名（兼容新旧两版字段名）
FIELD_MAP = {
    "品牌ID":       "brand_id",
    "品牌名称":     "brand_name",
    "品类名称":     "category_name",
    "活动ID":       "activity_id",
    "活动名称":     "activity_name",
    "批次名称":     "batch_name",
    "券批次id":     "batch_id",
    "活动价格力（各渠道有值渠道的算术平均）": "price_power",
    "活动开始时间": "start_date",
    "活动结束时间": "end_date",
    "发券总库存":   "total_stock",
    "券剩余库存":   "remain_stock",
    # 新版字段名（25列版，带"最大值"后缀）
    "曝光数(最大值)":      "exposure_pv",
    "领取数(最大值)":      "claim_pv",
    "核销数(最大值)":      "redeem_pv",
    "曝光uin数(最大值)":   "exposure_uv",
    "领取uin数(最大值)":   "claim_uv",
    "核销uin数(最大值)":   "redeem_uv",
    "领取到店率_uv\t(最大值)":  "claim_to_store_rate_uv",
    "到店核销率_uv\t(最大值)":  "store_redeem_rate_uv",
    "核销到店率_uv\t(最大值)":  "redeem_to_store_rate_uv",
    "到店未达门槛占比(最大值)":  "store_below_threshold",
    # 旧版字段名（19列版，无后缀）
    "当日曝光次数": "exposure_pv",
    "当日领取次数": "claim_pv",
    "当日核销次数": "redeem_pv",
    "当日曝光用户数": "exposure_uv",
    "当日领取用户数": "claim_uv",
    "当日核销用户数": "redeem_uv",
    "近7日均_到店核销率_uv":    "store_redeem_rate_uv",
    "近7日均_到店未达门槛占比": "store_below_threshold",
    # 新增券信息
    "优惠门槛":     "threshold_amount",
    "优惠金额":     "discount_amount",
    # 限领字段（iWiki CSV 新增的 P/Q 列；腾讯文档老底表没有此列）
    "单用户限领":   "single_user_limit",
    "单日限领":     "daily_limit",
}

# 备选字段名（底表可能有不同的列名格式）
FIELD_ALIASES = {
    "品牌id": "brand_id",
    "活动id": "activity_id",
    # 新 doc 4020370034 的 CSV 中文列名（旧 doc 与之对应的名字见 FIELD_MAP）
    "批次id": "batch_id",           # 旧 doc 老格式: 券批次id
    "投放开始时间": "start_date",    # 旧 doc 老格式: 活动开始时间
    "投放结束时间": "end_date",      # 旧 doc 老格式: 活动结束时间
    "活动价格力": "price_power",
    # 旧 doc 2026-07-06 起的中文正式命名（简化列名，含单位后缀）
    "批次ID":          "batch_id",
    "品类":            "category_name",
    "活动开始日期":    "start_date",
    "活动结束日期":    "end_date",
    "优惠门槛(分)":    "threshold_amount",
    "优惠金额(分)":    "discount_amount",
    "单用户领取上限":  "single_user_limit",
    "单日领取上限":    "daily_limit",
    "曝光数":          "exposure_pv",
    "领取数":          "claim_pv",
    "核销数":          "redeem_pv",
    "曝光用户数":      "exposure_uv",
    "领取用户数":      "claim_uv",
    "核销用户数":      "redeem_uv",
    # 兜底 doc(4020370034) 裸表头（无"最大值"/"当日"前缀）
    "曝光uin数":       "exposure_uv",
    "领取uin数":       "claim_uv",
    "核销uin数":       "redeem_uv",
    "领取到店率":      "claim_to_store_rate_uv",
    "到店核销率":      "store_redeem_rate_uv",
    "核销到店率":      "redeem_to_store_rate_uv",
    "到店未达门槛占比": "store_below_threshold",
    "近7日均_到店核销率UV": "store_redeem_rate_uv",
    "近7日均_到店核销率_UV": "store_redeem_rate_uv",
    "近7日均_到店未达门槛占比_uv": "store_below_threshold",
    "领取到店率_uv(最大值)": "claim_to_store_rate_uv",
    "到店核销率_uv(最大值)": "store_redeem_rate_uv",
    "核销到店率_uv(最大值)": "redeem_to_store_rate_uv",
}


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _safe_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(",", "").replace(" ", "")))
    except (ValueError, TypeError):
        return 0


def _safe_numeric(val):
    """安全转数值，保留小数"""
    if val is None:
        return None
    s = str(val).replace(",", "").replace(" ", "").replace("%", "")
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _extract_report_date(text: str) -> str:
    """
    从导览表文本提取 report_date
    格式1: "0420" → "2026-04-20"
    格式2: "20260420" → "2026-04-20"
    格式3: "0414-0420" → "2026-04-20" (取结束日期)
    格式4: "20260414_xxx" → "2026-04-14"
    """
    text = text.strip()

    # 格式3: "0414-0420" 取结束日期
    m = re.search(r'(\d{2})(\d{2})-(\d{2})(\d{2})', text)
    if m:
        month, day = m.group(3), m.group(4)
        return f"2026-{month}-{day}"

    # 格式2/4: "20260420..." 完整8位日期
    m = re.match(r'(\d{4})(\d{2})(\d{2})', text)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

    # 格式1: "0420" or "0420_xxx" 开头4位日期
    m = re.match(r'^(\d{2})(\d{2})', text)
    if m:
        return f"2026-{m.group(1)}-{m.group(2)}"

    return ""


def _map_row(row: dict, report_date: str):
    """将底表一行映射为 DB record"""
    record = {"report_date": report_date}

    # 先用主字段映射（精确匹配）
    for src, db_col in FIELD_MAP.items():
        if src in row:
            record[db_col] = row[src]

    # 再用别名补充缺失字段
    for src, db_col in FIELD_ALIASES.items():
        if db_col not in record or record[db_col] is None:
            if src in row:
                record[db_col] = row[src]

    # 模糊匹配：底表字段名可能带tab或空格，做 contains 匹配
    for row_key, row_val in row.items():
        if row_val is None:
            continue
        rk = row_key.replace('\t', '').replace(' ', '').strip()
        if '领取到店率' in rk and 'claim_to_store_rate_uv' not in record:
            record['claim_to_store_rate_uv'] = row_val
        elif '到店核销率' in rk and rk.startswith('到店') and 'store_redeem_rate_uv' not in record:
            record['store_redeem_rate_uv'] = row_val
        elif '核销到店率' in rk and 'redeem_to_store_rate_uv' not in record:
            record['redeem_to_store_rate_uv'] = row_val
        elif '到店未达门槛' in rk and 'store_below_threshold' not in record:
            record['store_below_threshold'] = row_val
        elif '券类型' in rk and 'coupon_type' not in record:
            record['coupon_type'] = row_val

    # 必须有 activity_id
    activity_id = _safe_str(record.get("activity_id", ""))
    if not activity_id:
        return None

    # 类型转换
    record["activity_id"] = activity_id
    record["brand_id"] = _safe_str(record.get("brand_id", ""))
    record["brand_name"] = _safe_str(record.get("brand_name", ""))
    record["category_name"] = _safe_str(record.get("category_name", ""))
    record["activity_name"] = _safe_str(record.get("activity_name", ""))
    record["batch_name"] = _safe_str(record.get("batch_name", ""))
    record["batch_id"] = _safe_str(record.get("batch_id", ""))
    record["coupon_type"] = _safe_str(record.get("coupon_type", ""))
    record["start_date"] = _safe_str(record.get("start_date", ""))
    record["end_date"] = _safe_str(record.get("end_date", ""))

    record["price_power"] = _safe_numeric(record.get("price_power"))
    record["threshold_amount"] = _safe_numeric(record.get("threshold_amount"))
    record["discount_amount"] = _safe_numeric(record.get("discount_amount"))
    record["total_stock"] = _safe_int(record.get("total_stock", 0))
    record["remain_stock"] = _safe_int(record.get("remain_stock", 0))

    record["exposure_pv"] = _safe_int(record.get("exposure_pv", 0))
    record["claim_pv"] = _safe_int(record.get("claim_pv", 0))
    record["redeem_pv"] = _safe_int(record.get("redeem_pv", 0))
    record["exposure_uv"] = _safe_int(record.get("exposure_uv", 0))
    record["claim_uv"] = _safe_int(record.get("claim_uv", 0))
    record["redeem_uv"] = _safe_int(record.get("redeem_uv", 0))

    record["store_redeem_rate_uv"] = _safe_numeric(record.get("store_redeem_rate_uv"))
    record["claim_to_store_rate_uv"] = _safe_numeric(record.get("claim_to_store_rate_uv"))
    record["redeem_to_store_rate_uv"] = _safe_numeric(record.get("redeem_to_store_rate_uv"))
    record["store_below_threshold"] = _safe_numeric(record.get("store_below_threshold"))

    # 限领字段：BIGINT，老底表没这两列时存 NULL（不要存 0，否则会被预警逻辑误判）
    raw_user_limit = record.get("single_user_limit")
    raw_daily_limit = record.get("daily_limit")
    record["single_user_limit"] = _safe_int(raw_user_limit) if raw_user_limit not in (None, "") else None
    record["daily_limit"]       = _safe_int(raw_daily_limit) if raw_daily_limit not in (None, "") else None

    return record


def _is_iwiki_date(report_date: str) -> bool:
    """判断给定 report_date (YYYY-MM-DD) 是否走 iWiki 源"""
    try:
        d = datetime.strptime(report_date, "%Y-%m-%d").date()
        return d >= IWIKI_CUTOFF_DATE
    except (ValueError, TypeError):
        return False


def _sync_one_date_from_iwiki(client, report_date: str, attachments: dict) -> int:
    """
    从 iWiki 拉单天数据并写库，返回写入条数。
    """
    print(f"\n  [iWiki 源] 处理: {report_date}")
    try:
        rows = load_activity_daily_from_iwiki(report_date, attachments)
    except Exception as e:
        print(f"    ✗ 加载失败: {e}")
        return 0

    if not rows:
        print("    ⚠ CSV 无数据，跳过")
        return 0

    sample_keys = list(rows[0].keys())[:8]
    print(f"    字段预览: {sample_keys}")

    records = []
    for row in rows:
        rec = _map_row(row, report_date)
        if rec:
            records.append(rec)
    print(f"    有效记录: {len(records)} 条")

    if records:
        upsert_batch(
            client, "tem_activity_daily", records,
            conflict_columns=["activity_id", "report_date"],
        )
    return len(records)


def _sync_one_date_from_doc(client, file_id: str, report_date: str, label: str) -> int:
    """
    从腾讯文档拉单天数据并写库，返回写入条数。
    """
    print(f"\n  [腾讯文档源] 处理: {label} → date={report_date} (file_id={file_id})")
    try:
        rows = read_target_sheet(file_id, header_row=1, data_start_row=2)
        print(f"    获取到 {len(rows)} 行, {len(rows[0]) if rows else 0} 列")

        if rows:
            sample_keys = list(rows[0].keys())[:8]
            print(f"    字段预览: {sample_keys}")

        records = []
        for row in rows:
            rec = _map_row(row, report_date)
            if rec:
                records.append(rec)
        print(f"    有效记录: {len(records)} 条")

        if records:
            upsert_batch(
                client, "tem_activity_daily", records,
                conflict_columns=["activity_id", "report_date"],
            )
        return len(records)
    except Exception as e:
        print(f"    ✗ 处理失败: {e}")
        import traceback
        traceback.print_exc()
        return 0


def sync_v2(latest_only: bool = False, date_filter: str = ""):
    """
    V2 同步主流程
    路由规则：
      - report_date >= 2026-05-26 → iWiki CSV
      - report_date <  2026-05-26 → 腾讯文档导览表
    两边都会按 latest_only / date_filter 收敛要处理的日期。
    """
    print("\n" + "=" * 60)
    print("V2 同步 tem_activity_daily（每日活动快照·混合源）")
    print(f"切换日期：>= {IWIKI_CUTOFF_DATE} 走 iWiki，否则走腾讯文档")
    print("=" * 60)

    client = get_client()
    total_written = 0

    # ── 1. iWiki 数据源（>= 2026-05-26）──
    iwiki_processed_dates: set = set()
    try:
        iwiki_atts = load_iwiki_activity_attachments()
    except Exception as e:
        print(f"\n[警告] iWiki 附件列表加载失败：{e}\n  → 跳过 iWiki 源，仅同步腾讯文档源")
        iwiki_atts = {}

    if iwiki_atts:
        iwiki_dates = sorted(iwiki_atts.keys())  # YYYY-MM-DD 升序
        # 只取 cutoff 之后的
        iwiki_dates = [d for d in iwiki_dates if _is_iwiki_date(d)]

        if latest_only and iwiki_dates:
            iwiki_dates = [iwiki_dates[-1]]
            print(f"\n[iWiki] --latest → 只处理 {iwiki_dates[0]}")
        elif date_filter:
            # 支持 "2026-05-26" 或 "0526" 两种写法
            normalized = _normalize_date_filter(date_filter)
            iwiki_dates = [d for d in iwiki_dates if d == normalized] if normalized \
                else [d for d in iwiki_dates if date_filter in d]
            print(f"\n[iWiki] 过滤 '{date_filter}' → {len(iwiki_dates)} 个日期")
        else:
            print(f"\n[iWiki] 待处理 {len(iwiki_dates)} 个日期: "
                  f"{iwiki_dates[0] if iwiki_dates else '(无)'} ~ "
                  f"{iwiki_dates[-1] if iwiki_dates else '(无)'}")

        for d in iwiki_dates:
            total_written += _sync_one_date_from_iwiki(client, d, iwiki_atts)
            iwiki_processed_dates.add(d)

    # ── 2. 腾讯文档源（< 2026-05-26）──
    # 如果 iWiki 路径已经处理过最新一天 / 命中过 date_filter，且用户用 --latest，
    # 就不必再走腾讯文档；否则继续按导览表做老路径。
    if latest_only and iwiki_processed_dates:
        # --latest 在 iWiki 已经处理过，老数据不再重跑
        pass
    else:
        try:
            doc_entries = _load_doc_nav_entries()
        except Exception as e:
            print(f"\n[腾讯文档] 导览表加载失败：{e}")
            doc_entries = []

        # 计算每条 entry 的 report_date，并过滤到 < cutoff 的
        for entry in doc_entries:
            entry["report_date"] = _extract_report_date(entry["text"])

        valid_doc = [
            e for e in doc_entries
            if e["file_id"] and e["url"] and e["report_date"]
            and not _is_iwiki_date(e["report_date"])
        ]
        print(f"\n[腾讯文档] 过滤后待处理 {len(valid_doc)} 条 (cutoff < {IWIKI_CUTOFF_DATE})")

        if not valid_doc:
            pass
        elif latest_only:
            # 历史最新的一条（按 report_date 倒序）
            valid_doc.sort(key=lambda e: e["report_date"], reverse=True)
            valid_doc = [valid_doc[0]]
            print(f"  --latest → 只处理 {valid_doc[0]['text']}")
        elif date_filter:
            filtered = [e for e in valid_doc if date_filter in e["text"]]
            if not filtered:
                print(f"  ⚠ 未找到包含 '{date_filter}' 的条目")
                valid_doc = []
            else:
                valid_doc = filtered
                print(f"  → 匹配到 {len(valid_doc)} 条")

        for entry in valid_doc:
            total_written += _sync_one_date_from_doc(
                client,
                file_id=entry["file_id"],
                report_date=entry["report_date"],
                label=entry["text"],
            )

    print(f"\n{'=' * 60}")
    print(f"V2 同步完成！共写入 {total_written} 条记录")
    print(f"{'=' * 60}")


def _normalize_date_filter(s: str) -> str:
    """把 '0526' 或 '2026-05-26' 统一成 'YYYY-MM-DD'，无法识别返回空"""
    s = (s or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    if re.fullmatch(r"\d{4}", s):  # MMDD
        return f"{datetime.now().year}-{s[:2]}-{s[2:]}"
    if re.fullmatch(r"\d{8}", s):  # YYYYMMDD
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return ""


def _load_doc_nav_entries() -> list:
    """读取腾讯文档导览表，返回 [{text, url, file_id}, ...]"""
    url = export_file(NAV_FILE_ID)
    xlsx_path = download_xlsx(url)
    hyperlinks = extract_hyperlinks(xlsx_path)

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]
    entries = []
    for row_num in range(1, (ws.max_row or 100) + 1):
        text = ws.cell(row_num, 1).value
        if not text:
            continue
        text = str(text).strip()
        link = hyperlinks.get(f"A{row_num}", "")
        fid = extract_file_id_from_url(link) if link else ""
        entries.append({"text": text, "url": link, "file_id": fid})
    wb.close()
    os.unlink(xlsx_path)
    return entries


def main():
    parser = argparse.ArgumentParser(description="V2 每日活动快照同步")
    parser.add_argument("--latest", action="store_true",
                        help="只同步最新一天")
    parser.add_argument("--date", type=str, default="",
                        help="只同步指定日期（匹配导览文本，如 '0420'）")
    args = parser.parse_args()

    sync_v2(latest_only=args.latest, date_filter=args.date)


if __name__ == "__main__":
    main()
