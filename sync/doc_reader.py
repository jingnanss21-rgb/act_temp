"""
doc_reader.py - 腾讯文档读取器
封装 MCP export + xlsx 解析的完整流程
"""
import os
import time
import json
import zipfile
import tempfile
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

MCP_URL = "https://docs.qq.com/openapi/mcp"


def _get_headers():
    token = os.getenv("TENCENT_DOCS_TOKEN", "")
    return {
        "Authorization": token,
        "Content-Type": "application/json",
    }


def _mcp_call(tool_name: str, arguments: dict) -> dict:
    """通用 MCP 工具调用"""
    payload = {
        "jsonrpc": "2.0",
        "id": 1,
        "method": "tools/call",
        "params": {
            "name": tool_name,
            "arguments": arguments,
        },
    }
    resp = requests.post(MCP_URL, headers=_get_headers(), json=payload, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    if "error" in data and data["error"]:
        raise RuntimeError(f"MCP error: {data['error']}")
    # MCP 返回格式：result.content[0].text (JSON string)
    result = data.get("result", {})
    content_list = result.get("content", [])
    if content_list:
        text = content_list[0].get("text", "{}")
        return json.loads(text)
    return result


def export_file(file_id: str) -> str:
    """导出文档为 xlsx，返回下载 URL"""
    print(f"  → 导出文档 {file_id} ...")
    result = _mcp_call("manage.export_file", {"file_id": file_id})
    task_id = result.get("task_id", "")
    if not task_id:
        raise RuntimeError(f"export_file 未返回 task_id: {result}")

    # 轮询导出进度
    for _ in range(30):  # 最多等 150 秒
        time.sleep(5)
        progress = _mcp_call("manage.export_progress", {"task_id": task_id})
        p = progress.get("progress", 0)
        print(f"    导出进度: {p}%")
        if p >= 100:
            file_url = progress.get("file_url", "")
            if file_url:
                return file_url
            raise RuntimeError(f"导出完成但无下载链接: {progress}")
    raise RuntimeError("导出超时")


def download_xlsx(url: str, save_path: str = None) -> str:
    """下载 xlsx 文件到本地"""
    if save_path is None:
        fd, save_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
    print(f"  → 下载到 {save_path}")
    resp = requests.get(url, timeout=120)
    resp.raise_for_status()
    with open(save_path, "wb") as f:
        f.write(resp.content)
    return save_path


def parse_xlsx_rows(xlsx_path: str, sheet_index: int = 0) -> list[dict]:
    """
    解析 xlsx 文件，返回 [{header: value, ...}, ...]
    第一行为表头
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_index]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        result.append(record)
    return result


def extract_hyperlinks(xlsx_path: str, sheet_name: str = "sheet1") -> dict:
    """
    从 xlsx 文件中提取超链接
    返回 {cell_ref: url, ...}，如 {"A2": "https://docs.qq.com/..."}
    """
    hyperlinks = {}
    try:
        zf = zipfile.ZipFile(xlsx_path)
        # 尝试多个可能的 rels 文件路径
        rels_paths = [
            "xl/worksheets/_rels/sheet1.xml.rels",
            "xl/worksheets/_rels/sheet2.xml.rels",
        ]
        for rels_path in rels_paths:
            try:
                rels_tree = ET.parse(zf.open(rels_path))
            except KeyError:
                continue

            rels_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            links = {}
            for rel in rels_tree.findall(".//r:Relationship", rels_ns):
                if "hyperlink" in rel.get("Type", "").lower():
                    links[rel.get("Id")] = rel.get("Target")

            if not links:
                continue

            # 解析对应的 sheet xml
            sheet_xml_path = rels_path.replace("_rels/", "").replace(".xml.rels", ".xml")
            ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            sheet_tree = ET.parse(zf.open(sheet_xml_path))
            for hl in sheet_tree.findall(".//s:hyperlinks/s:hyperlink", ns):
                ref = hl.get("ref")
                rid = hl.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                )
                if rid and rid in links:
                    hyperlinks[ref] = links[rid]
            if hyperlinks:
                break

        zf.close()
    except Exception as e:
        print(f"  ⚠ 提取超链接失败: {e}")
    return hyperlinks


def extract_file_id_from_url(url: str) -> str:
    """
    从腾讯文档 URL 提取 file_id
    例：https://docs.qq.com/sheet/DWUF3Wk1wYnNUUVhE?tab=BB08J2 → DWUF3Wk1wYnNUUVhE
    例：https://docs.qq.com/smartsheet/DWEhnUHdDVEJ1Y05i?from_wiki_space=1 → DWEhnUHdDVEJ1Y05i
    """
    from urllib.parse import urlparse
    parsed = urlparse(url)
    parts = parsed.path.strip("/").split("/")
    if len(parts) >= 2:
        return parts[-1]
    return ""


def read_index_table_with_links(file_id: str) -> list[dict]:
    """
    读取索引表（表1或表5），返回 [{text: ..., url: ...}, ...]
    索引表的特点：每行有一个文本+超链接，指向目标文档
    """
    url = export_file(file_id)
    xlsx_path = download_xlsx(url)
    hyperlinks = extract_hyperlinks(xlsx_path)
    rows = parse_xlsx_rows(xlsx_path)

    # 把超链接关联到行
    result = []
    for i, row in enumerate(rows):
        cell_ref = f"A{i + 2}"  # A2 开始（A1 是表头）
        link = hyperlinks.get(cell_ref, "")
        first_val = list(row.values())[0] if row else ""
        result.append({
            "text": str(first_val) if first_val else "",
            "url": link,
            "file_id": extract_file_id_from_url(link) if link else "",
        })

    os.unlink(xlsx_path)
    return result


def read_target_sheet(file_id: str) -> list[dict]:
    """
    读取目标文档（普通 sheet 或 smartsheet 导出后的 xlsx），
    返回解析后的行数据
    """
    url = export_file(file_id)
    xlsx_path = download_xlsx(url)
    rows = parse_xlsx_rows(xlsx_path)
    os.unlink(xlsx_path)
    return rows
