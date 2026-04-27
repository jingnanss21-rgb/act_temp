"""
sync_pinned.py - V3 策略分析数据同步
同步两张底表：
  1. 置顶操作记录 → tem_pinned_ops
  2. 腰部达标标记 → tem_waist_qualified

不影响现有的 sync_all.py 和 sync_v2.py。

用法:
  python3 sync_pinned.py            # 同步两张表
  python3 sync_pinned.py --table pinned_ops
  python3 sync_pinned.py --table waist_qualified
"""
import os
import re
import argparse
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

from doc_reader import export_file, download_xlsx
from supabase_writer import get_client, upsert_batch
from openpyxl import load_workbook


# 文档ID
DOC_IDS = {
    "pinned_ops":      "DWW1oUktoeUVzUGx2",  # 置顶操作记录
    "waist_qualified": "DWXV0R3dNekVGS1hk",  # 腰部达标标记
}


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _safe_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def _parse_pin_date(val):
    """
    解析置顶时间为 YYYY-MM-DD
    支持:
      - datetime对象
      - "2026-04-14 16:02:36"
      - "2026-04-14"
    """
    if val is None:
        return None
    # datetime 对象
    if hasattr(val, 'strftime'):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # 匹配前10位 YYYY-MM-DD
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def sync_pinned_ops(client):
    """
    同步置顶操作记录
    底表字段:
      A: id
      B: row_creator_58
      C: brand_id_51  ← 品牌ID
      D: strategy_id_72
      E: strategy_act_id_96
      F: row_createtime_87  ← 创建时间
      G: row_modifytime_14  ← 修改时间（业务定义为置顶时间）
    一品牌一条记录
    """
    print("\n" + "=" * 60)
    print("同步 tem_pinned_ops（置顶操作记录）")
    print("=" * 60)

    url = export_file(DOC_IDS["pinned_ops"])
    xlsx_path = download_xlsx(url)
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    records = []
    # 第一行是表头（英文字段名），数据从第2行起
    for row_num in range(2, (ws.max_row or 0) + 1):
        brand_id = _safe_str(ws.cell(row_num, 3).value)  # C列
        pin_time = ws.cell(row_num, 7).value              # G列
        pin_date = _parse_pin_date(pin_time)

        if not brand_id or not pin_date:
            continue

        records.append({
            "brand_id":     brand_id,
            "brand_name":   "",  # 底表无品牌名，后续前端从brand_daily关联
            "pin_date":     pin_date,
            "row_creator":  _safe_str(ws.cell(row_num, 2).value),
            "strategy_id":  _safe_str(ws.cell(row_num, 4).value),
        })

    wb.close()
    os.unlink(xlsx_path)

    print(f"  有效记录: {len(records)} 条")

    # 同一brand_id取最新pin_date
    best = {}
    for r in records:
        bid = r["brand_id"]
        if bid not in best or r["pin_date"] > best[bid]["pin_date"]:
            best[bid] = r
    dedup_records = list(best.values())
    print(f"  去重后: {len(dedup_records)} 条")

    if dedup_records:
        upsert_batch(client, "tem_pinned_ops", dedup_records,
                      conflict_columns=["brand_id"])


def sync_waist_qualified(client):
    """
    同步腰部达标标记
    底表字段:
      A: 品牌ID
      B: 品牌名称
      C: 服务商
      D: 类目
      E: 四要素是否达标  ← 达标/未达标
      F: 达标分类
      G: 人工打标具体反馈
      H: 是否存活
      I: 7日老客应曝UV（可选）
    """
    print("\n" + "=" * 60)
    print("同步 tem_waist_qualified（腰部达标）")
    print("=" * 60)

    url = export_file(DOC_IDS["waist_qualified"])
    xlsx_path = download_xlsx(url)
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    records = []
    for row_num in range(2, (ws.max_row or 0) + 1):
        brand_id = _safe_str(ws.cell(row_num, 1).value)
        if not brand_id:
            continue

        records.append({
            "brand_id":             brand_id,
            "brand_name":           _safe_str(ws.cell(row_num, 2).value),
            "operating_sp":         _safe_str(ws.cell(row_num, 3).value),
            "category":             _safe_str(ws.cell(row_num, 4).value),
            "is_qualified":         _safe_str(ws.cell(row_num, 5).value),
            "qualified_type":       _safe_str(ws.cell(row_num, 6).value),
            "manual_feedback":      _safe_str(ws.cell(row_num, 7).value),
            "is_alive":             _safe_str(ws.cell(row_num, 8).value),
            "old_user_should_uv_w7": _safe_int(ws.cell(row_num, 9).value),
        })

    wb.close()
    os.unlink(xlsx_path)

    # 去重（同brand_id可能有重复）
    best = {}
    for r in records:
        best[r["brand_id"]] = r
    dedup_records = list(best.values())
    print(f"  有效记录: {len(dedup_records)} 条（已去重）")

    if dedup_records:
        upsert_batch(client, "tem_waist_qualified", dedup_records,
                      conflict_columns=["brand_id"])


SYNC_MAP = {
    "pinned_ops": sync_pinned_ops,
    "waist_qualified": sync_waist_qualified,
}


def main():
    parser = argparse.ArgumentParser(description="V3 策略分析数据同步")
    parser.add_argument("--table", type=str, default="all",
                        choices=["all"] + list(SYNC_MAP.keys()),
                        help="指定要同步的表（默认 all）")
    args = parser.parse_args()

    client = get_client()
    print(f"Supabase 连接成功: {os.getenv('SUPABASE_URL')}")

    if args.table == "all":
        for name in ["pinned_ops", "waist_qualified"]:
            SYNC_MAP[name](client)
    else:
        SYNC_MAP[args.table](client)

    print("\n✓ V3 策略分析同步完成!")


if __name__ == "__main__":
    main()
